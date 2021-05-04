# Sofía Múnera Medina
# Python
# POO con archivos .xslx
# Fuentes: https://www.bancomundial.org/es/home


# EJEMPLO 1

# Importamos la libreria pandas para poder tratar archivos de Excel

import pandas as pd

# ExcelWriter lo usaremos para crear un nuevo archivo .xlsx

from pandas import ExcelWriter

# Importamos openpyxl para leer un archivo .xlsx

import openpyxl


# Creamos la clase pais

class Pais:

  # Documentacion 

  ''' Lee la informacion de un archivo .xls, la modifica y crea otro donde almacena los cambios '''

  # Mediante __init__ crearemos los atributos de instancia 

  def __init__(self):

    # Documentacion 

    ''' Metodo constructor de la clase Pais, establece 3 atributos de instancia con valores predefinidos '''

    # Creamos los atributos de instancia, sus valores serán cadenas de texto que representan 3 paises 

    self.pais0='España'
    self.pais1='Estados Unidos'
    self.pais2='China'
  
  # Creamos un metodo llamado de la clase pais llamado modificar 

  def modificar(self):

    # Documentacion 

    ''' Lee un archivo .xlsx, almacena la informacion que contiene y luego crea un nuevo archivo con esta informacion y otra adicional '''

    # Usando la libreria openpyxl y la funcion load_workbook vamos a leer el archivo que usaremos de base llamado Libro1.xlsx

    libro1=openpyxl.load_workbook('Libro1.xlsx')

    # Ahora vamos a usar la Hoja1 del libro que acabamos de leer 

    hoja1=libro1['Hoja1']

    # Mediante Hoja1[celdas] vamos a leer los datos que usaremos en el archivo nuevo. De A2:A4 hay una lista de 3 capitales y de B2:B4 estan los idiomas que hablan en cada una de estas ciudades

    capitales=hoja1['A2':'A4']
    idiomas=hoja1['B2':'B4']


    # Ahora comenzaremos a crear un nuevo libro

    # Creamos una lista con los atributos de instancia

    paises=[self.pais0,self.pais1,self.pais2]

    # Mediante DataFrame crearemos una tabla donde las cabeceras de cada columna son Pais,Capital e Idioma y a cada uno de estos encabezados le vamos a asignar unos datos. En el caso de Capital e Idioma los datos serán los que leimos del archivo original (Libro1) y para Pais se le signa la lista que contiene los valores de los atributos de instancia 

    copia1=pd.DataFrame({'Pais':paises,'Capital':capitales,'Idioma':idiomas})

    # Crearemos una nueva lista con informacion de continentes 

    continentes=['Europa','América','Asia']

    # Insertaremos una columna con encabezado Continente y valores de la lista continentes. Esta irá en la posicion/locacion 2, lo que significará que la columna Idioma se desplazará una a la derecha 

    # el orden de columnas será Pais, Capital, Continente y por ultimo Idioma

    copia1.insert(2,'Continente',continentes,allow_duplicates=False)

    # mediante Excel writer crearemos un archivo llamado Copia1.xlsx

    archivo=ExcelWriter('Copia1.xlsx')

    # Ahora con ayuda de .to_excel el Dataframe que creamos (incluyendo la informacion insertada posteriormente) lo guardaremos en este nuevo archivo (Copia1), la hoja se llamara Hoja Copia y podemos index=False para que no se incluya encabezado de numeros 

    copia1.to_excel(archivo,'Hoja Copia',index=False)

    # Guardamos lo ingresado al archivo copia 1

    archivo.save()

    

# Llamamos la clase Pais, que inmediantamente creará los 3 atributos de instancia 

a=Pais()

# Mediante el metodo modificar vamos a leer la información de Libro1 y crearemos un nuevo archivo llamado Copia 1 

a.modificar()

# FINALIZA EJEMPLO 1

"""
# EJEMPLO 2

# Creamos una clase llamada Pais

class Pais:

  # Documentacion 

  ''' Clase que representa la poblacion total y de mujeres de 3 paises diferentes '''

  # Usamos __init__ para establecer los atributos de instancia 

  def __init__(self):

    # Documentacion 

    ''' Metodo constructor de la clase Pais, establece dos atributos de instancia de tipo cadena de texto '''

    # Establecemos los dos atribtuos que luego usaremos como encabezados de una tabla 

    self.encabezado0='PAIS'
    self.encabezado1='POBLACIÓN TOTAL'
  
  # CReamos una funcion llamada agregar donde crearemos un nuevo archivo de excel y haremos uso de los atributos 

  def agregar(self):

    # Documentacion 

    ''' Lee un archivo llamado Libro2.xlsx, almacena la informacion de este y luego crea un nuevo archivo con la informacion extraida y una adicional '''

    # Usando la libreria openpyxl y la funcion load_workbook vamos a leer el archivo que usaremos de base llamado Libro2.xlsx

    libro1=openpyxl.load_workbook('Libro2.xlsx')

    # Ahora vamos a usar la Hoja1 del libro que acabamos de leer 

    hoja1=libro1['Hoja1']

    # Mediante Hoja1[celdas] vamos a leer los datos que usaremos en el archivo nuevo. De A2:A4 hay una lista de 3 paises y de B2:B4 esta la poblacion total de cada pais 

    paises=hoja1['A2':'A4']
    poblacion_total=hoja1['B2':'B4']

    # Creamos una nueva lista con el porcentaje de mujeres,migrantes, personas que viven en tugurios  en cada pais

    poblacion_mujeres=hoja1[50.9,48.0,51.1]
    poblacion_migrantes=hoja1[0.3,0.4,0.9]
    poblacion_tugurios=[28,35,16]

    # Mediante dataFrame vamos a crear una tabla con los datos ingresados donde los encabezados de las dos primeras columnas serán los atributos y las otras tres columnas tendran como titulo MUJERES,MIGRANTES Y POBLACION QUE VIVE EN TUGURIOS estas tres ultimas columnas tendran sus datos en porcentaje

    copia2=pd.DataFrame({self.encabezado0:paises,self.encabezado1:poblacion_total,'MUJERES(%)':poblacion_mujeres,' MIGRANTES(%)':poblacion_migrantes,'POBLACION QUE VIVE EN TUGURIOS (%)':poblacion_tugurios})

    import xlsxwriter 

    encabezado = copia2.add_format()
    encabezado.set_font_color('blue')
    encabezado.set_font_size(16)
    encabezado.set_bold()

    copia2_estadistica=hoja1.describe

    # Con ExcelWritel creamos una nueva hoja llamada copia2 de tipo xlsx

    nuevo=ExcelWriter('Copia2.xlsx')

    # Ahora con ayuda de .to_excel el Dataframe que creamos lo guardaremos en este nuevo archivo (Copia2), la hoja se llamara Hoja Copia 2 y podemos index=False para que no se incluya encabezado de numeros 

    copia2.to_excel(nuevo,'Hoja Copia 2',index=False)
    copia2_estadistica.to_excel(nuevo,'Hoja Copia 2.2',index=False)


    # Guardamos lo ingresado al archivo copia 2

    nuevo.save()

    

b=Pais()
b.agregar

# FINALIZA EJEMPLO 2
"""

class ejemplo:

  def __init__(self):

    self.raiz='raiz cuadrada'
    self.sum='SUMATORIA'

  def nuevo(self):

    tabla=pd.read_excel('Libro3.xlsx', 'Hoja1')

    import numpy as np

    hoja0=tabla.apply(np.sqrt)

    datos=tabla.apply(np.sum)

    hoja1=pd.DataFrame({self.sum:datos})

    nuevo=ExcelWriter('Copia3.xlsx')

    # Ahora con ayuda de .to_excel el Dataframe que creamos lo guardaremos en este nuevo archivo (Copia2), la hoja se llamara Hoja Copia 2 y podemos index=False para que no se incluya encabezado de numeros 

    hoja0.to_excel(nuevo,'Raices',index=False)
    hoja1.to_excel(nuevo,'Sumatoria',index=False)


    # Guardamos lo ingresado al archivo copia 2

    nuevo.save()
  

a=ejemplo()
a.nuevo()












